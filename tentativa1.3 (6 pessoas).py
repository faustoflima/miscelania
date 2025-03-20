# from numpy import number
import pyautogui
import mouse
import random
from time import strftime, sleep
from PIL import Image
import pytesseract
import keyboard
import os
import holdem_calc
import re
import numpy as np
import cv2
from pynput.mouse import Button, Controller
# from sympy import true
import openpyxl
mousee = Controller()
pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract'

sessao = strftime("%y_%m_%d %H_%M_%S")
filepath = 'C:\\Users\Pichau\Documents\py4e\Poker/sessao ' + sessao
os.mkdir(filepath)

def fazer_screenshot(data_hora, filepath):
    screenshot = pyautogui.screenshot()
    screenshot.save(filepath + '/' +  data_hora + ".jpg")

posicao_desistir = (990, 990)
posicao_mesa = (1290, 990)
posicao_apostar = (1625, 990)
posicao_centralizar = (1319, 675)

x = 260
y = 50

def desistir():
    # descontinuada após que eu coloquei clique no teclado em vez de na tela
    # duracao = random.randint(2,5)/10
    # posicao_x = posicao_desistir[0] + random.randint(0,x)
    # posicao_y = posicao_desistir[1] + random.randint(0,y)

    # mouse.move(posicao_x, posicao_y, absolute=True, duration=duracao)
    # # pyautogui.click()
    # mousee.press(Button.left)
    # mousee.release(Button.left)

    keyboard.press('b')
    keyboard.release('b')

    print('acao desistir')
    # f.write('\nacao desistir')
    excel.cell(row = linha, column = 32).value = 'desistir' 
    wb.save(filepath+"\log.xlsx")


def mesa():
    # descontinuada após que eu coloquei clique no teclado em vez de na tela
    # duracao = random.randint(2,5)/10
    # posicao_x = posicao_mesa[0] + random.randint(0,x)
    # posicao_y = posicao_mesa[1] + random.randint(0,y)

    # mouse.move(posicao_x, posicao_y, absolute=True, duration=duracao)
    # # pyautogui.click()
    # mousee.press(Button.left)
    # mousee.release(Button.left)
    
    keyboard.press('n')
    keyboard.release('n')

    print('acao mesa/call')
    # f.write('\nacao mesa/call')
    excel.cell(row = linha, column = 32).value = 'mesa' 
    wb.save(filepath+"\log.xlsx")

def apostar():
    # descontinuada após que eu coloquei clique no teclado em vez de na tela
    # duracao = random.randint(2,5)/10
    # posicao_x = posicao_apostar[0] + random.randint(0,x)
    # posicao_y = posicao_apostar[1] + random.randint(0,y)

    # mouse.move(posicao_x, posicao_y, absolute=True, duration=duracao)
    # # pyautogui.click()
    # mousee.press(Button.left)
    # mousee.release(Button.left)

    keyboard.press('m')
    keyboard.release('m')

    print('acao apostar')
    # f.write('\nacao apostar')
    excel.cell(row = linha, column = 32).value = 'apostar' 
    wb.save(filepath+"\log.xlsx")


# descontinuada após que eu coloquei clique no teclado em vez de na tela
# def centralizar_mouse():
#     duracao = random.randint(1,4)/10
#     posicao_x = posicao_centralizar[0] + random.randint(0,x)
#     posicao_y = posicao_centralizar[1] + random.randint(0,y)

#     mouse.move(posicao_x, posicao_y, absolute=True, duration=duracao)

minha_carta_1 = (897, 670, 955, 752)
minha_carta_2 = (995, 667, 1052, 743)

minha_carta_1_naipe = (866, 710, 894, 738)
minha_carta_2_naipe = (964, 710, 992, 738)

carta_mesa_1 = (735, 371, 797, 445)
carta_mesa_2 = (839, 371, 900, 445)
carta_mesa_3 = (944, 371, 1004, 445)
carta_mesa_4 = (1048, 371, 1109, 445)
carta_mesa_5 = (1150, 371, 1213, 445)

carta_mesa_1_naipe = (708, 410, 735, 439)
carta_mesa_2_naipe = (811, 410, 839, 439)
carta_mesa_3_naipe = (914, 410, 942, 439)
carta_mesa_4_naipe = (1018, 410, 1047, 439)
carta_mesa_5_naipe = (1122, 410, 1149, 439)

cartas_do_baralho = ('A','K','Q','J','T','9','8','7','6','5','4','3','2')
naipes_do_baralho = ('h','s','c','d')
todas_cartas_do_baralho = []

for carta in cartas_do_baralho:
    for naipe in naipes_do_baralho:
        todas_cartas_do_baralho.append(carta+naipe)

# image = Image.open("Poker\screen_perfeita.jpg")
# text = pytesseract.image_to_string(image)
# print(text)

image_banco_A = Image.open("C:\\Users\Pichau\Documents\py4e\Poker/números\A.png")
image_banco_2 = Image.open("C:\\Users\Pichau\Documents\py4e\Poker/números/2.png")
image_banco_3 = Image.open("C:\\Users\Pichau\Documents\py4e\Poker/números/3.png")
image_banco_4 = Image.open("C:\\Users\Pichau\Documents\py4e\Poker/números/4.png")
image_banco_5 = Image.open("C:\\Users\Pichau\Documents\py4e\Poker/números/5.png")
image_banco_6 = Image.open("C:\\Users\Pichau\Documents\py4e\Poker/números/6.png")
image_banco_7 = Image.open("C:\\Users\Pichau\Documents\py4e\Poker/números/7.png")
image_banco_8 = Image.open("C:\\Users\Pichau\Documents\py4e\Poker/números\8.png")
image_banco_9 = Image.open("C:\\Users\Pichau\Documents\py4e\Poker/números\9.png")
image_banco_10 = Image.open("C:\\Users\Pichau\Documents\py4e\Poker/números/10.png")
image_banco_J = Image.open("C:\\Users\Pichau\Documents\py4e\Poker/números\J.png")
image_banco_Q = Image.open("C:\\Users\Pichau\Documents\py4e\Poker/números\Q.png")
image_banco_K = Image.open("C:\\Users\Pichau\Documents\py4e\Poker/números\K.png")

image_banco_C = Image.open("C:\\Users\Pichau\Documents\py4e\Poker/números/naipe_copas.png")
image_banco_E = Image.open("C:\\Users\Pichau\Documents\py4e\Poker/números/naipe_espadas.png")
image_banco_O = Image.open("C:\\Users\Pichau\Documents\py4e\Poker/números/naipe_ouro.png")
image_banco_P = Image.open("C:\\Users\Pichau\Documents\py4e\Poker/números/naipe_paus.png")

image_banco_dealer = Image.open("C:\\Users\Pichau\Documents\py4e\Poker/números/dealer.png")

image_banco_carta_adversario = Image.open("C:\\Users\Pichau\Documents\py4e\Poker/números/carta_adversario.jpg")

image_oponentes_mesa = Image.open('C:\\Users\Pichau\Documents\py4e\Poker/números\Estrela Oponente Ativo.jpg')

image_oponente_ausente = Image.open('C:\\Users\Pichau\Documents\py4e\Poker/números\Oponente Ausente.jpg')

def reconhecer_carta(imagem):
    if pyautogui.locate(image_banco_A, imagem, grayscale=True, confidence = 0.95) != None:
        return ('A')
        # print('A')
    elif pyautogui.locate(image_banco_2, imagem, grayscale=True, confidence = 0.95) != None:
        return ('2')
        # print('2')
    elif pyautogui.locate(image_banco_3, imagem, grayscale=True, confidence = 0.95) != None:
        return ('3')
        # print('3')
    elif pyautogui.locate(image_banco_4, imagem, grayscale=True, confidence = 0.95) != None:
        return ('4')
        # print('4')
    elif pyautogui.locate(image_banco_5, imagem, grayscale=True, confidence = 0.95) != None:
        return ('5')
        # print('5')
    elif pyautogui.locate(image_banco_6, imagem, grayscale=True, confidence = 0.95) != None:
        return ('6')
        # print('6')
    elif pyautogui.locate(image_banco_7, imagem, grayscale=True, confidence = 0.95) != None:
        return ('7')
        # print('7')
    elif pyautogui.locate(image_banco_8, imagem, grayscale=True, confidence = 0.95) != None:
        return ('8')
        # print('8')
    elif pyautogui.locate(image_banco_9, imagem, grayscale=True, confidence = 0.95) != None:
        return ('9')
        # print('9')
    elif pyautogui.locate(image_banco_10, imagem, grayscale=True, confidence = 0.90) != None:
        return ('T')
        # print('10')
    elif pyautogui.locate(image_banco_J, imagem, grayscale=True, confidence = 0.90) != None:
        return ('J')
        # print('J')
    elif pyautogui.locate(image_banco_Q, imagem, grayscale=True, confidence = 0.90) != None:
        return ('Q')
        # print('Q')
    elif pyautogui.locate(image_banco_K, imagem, grayscale=True, confidence = 0.90) != None:
        return ('K')
        # print('K')
    else:
        return (0)
        # print('Não Encontrado')

def reconhecer_carta_naipe(imagem):
    if pyautogui.locate(image_banco_C, imagem, grayscale=True, confidence = 0.8) != None:
        return ('h') #hearts ou coração
        # print('C')
    if pyautogui.locate(image_banco_E, imagem, grayscale=True, confidence = 0.8) != None:
        return ('s') #spades ou espadas
        # print('E')
    if pyautogui.locate(image_banco_O, imagem, grayscale=True, confidence = 0.8) != None:
        return ('d') #diamond ou ouro
        # print('O')
    if pyautogui.locate(image_banco_P, imagem, grayscale=True, confidence = 0.8) != None:
        return ('c') #clubs ou paus
        # print('P')
    else:
        return (0)
        # print('Não Encontrado')

def reconhecer_dealer(imagem):
    if pyautogui.locate(image_banco_dealer, imagem, grayscale=True, confidence = 0.90) == None:
        return (0)
    elif pyautogui.locate(image_banco_dealer, imagem, grayscale=True, confidence = 0.90)[0:2] == (792, 656):
        return (1)
    elif pyautogui.locate(image_banco_dealer, imagem, grayscale=True, confidence = 0.90)[0:2] == (487, 492):
        return (2)
    elif pyautogui.locate(image_banco_dealer, imagem, grayscale=True, confidence = 0.90)[0:2] == (511, 356):
        return (3)
    elif pyautogui.locate(image_banco_dealer, imagem, grayscale=True, confidence = 0.90)[0:2] == (1004, 262):
        return (4)
    elif pyautogui.locate(image_banco_dealer, imagem, grayscale=True, confidence = 0.90)[0:2] == (1378, 349):
        return (5)
    elif pyautogui.locate(image_banco_dealer, imagem, grayscale=True, confidence = 0.90)[0:2] == (1303, 600):
        return (6)
    # elif pyautogui.locate(image_banco_dealer, imagem, grayscale=True, confidence = 0.90)[0:2] == (1345, 331):
    #     return (7)
    # elif pyautogui.locate(image_banco_dealer, imagem, grayscale=True, confidence = 0.90)[0:2] == (1377, 531):
    #     return (8)
    # elif pyautogui.locate(image_banco_dealer, imagem, grayscale=True, confidence = 0.90)[0:2] == (1217, 610):
    #     return (9)
    else:
        return(0)

def fazer_acao (imagem):
    soma = 0

    if imagem.getpixel((986, 992)) == (191, 35, 36):
        soma = soma + 1
    if imagem.getpixel((1305, 992)) == (191, 35, 36):
        soma = soma + 2
    if imagem.getpixel((1618, 992)) == (191, 35, 36):
        soma = soma + 4
    return(soma)

def lista_apostas_oponentes (imagem):
    lista_oponentes = [2,3,4,5,6]

    if imagem.getpixel((589, 570)) == (15, 90, 34): #significa que não fez a aposta
        lista_oponentes.remove(2)
    if imagem.getpixel((611, 341)) == (13, 91, 33):
        lista_oponentes.remove(3)
    if imagem.getpixel((1036, 238)) == (15, 79, 29):
        lista_oponentes.remove(4)
    if imagem.getpixel((1290, 304)) == (12, 87, 31):
        lista_oponentes.remove(5)
    if imagem.getpixel((1329, 567)) == (13, 91, 33):
        lista_oponentes.remove(6)

    return(lista_oponentes)

def reconhecer_estagio_jogo (minha_carta1, minha_carta2, mesa_carta1, mesa_carta2, mesa_carta3, mesa_carta4, mesa_carta5):
    if minha_carta1 == 0 and minha_carta2 == 0:
        return('sem cartas')
    elif mesa_carta5 != 0:
        return ('river')
    elif mesa_carta4 != 0:
        return ('turn')
    elif mesa_carta3 != 0 and mesa_carta2 != 0 and mesa_carta1 != 0:
        return ('flop')
    elif minha_carta1 != 0 and minha_carta2 != 0:
        return ('pre-flop')
    else:
        return ('sem cartas')

def calcular_chance_de_ganhar(carta1, carta2, carta_mesa1, carta_mesa2, carta_mesa3, carta_mesa4, carta_mesa5, acao, num_oponentes):
    # odds = holdem_calc.calculate(board, exact_calculation, num_sims, read_from_file , hero_hand, verbose)

    #retirado temporariamente
    # if acao == 0:
    #     n = 20000
    #     print('calculo grande')
    # else:
    #     n = 5000
    #     print('calculo pequeno')

    # n = 20.000 --> 2,45s de processamento
    # n = 10.000 --> 1,40s de processamento
    # n = 5.000 --> 1,00s de processamento

    n = 10000
    
    i = 0
    soma_ganhar = 0
    soma_empatar = 0
    while i < n:
        cartas_baralho_restante = todas_cartas_do_baralho.copy()
        cartas_cheias = []
        cartas_vazias = []

        dicionario = {  'carta1': carta1, 
                        'carta2': carta2, 
                        'carta_mesa1': carta_mesa1,
                        'carta_mesa2': carta_mesa2,
                        'carta_mesa3': carta_mesa3,
                        'carta_mesa4': carta_mesa4,
                        'carta_mesa5': carta_mesa5}

        for k,v in dicionario.items():
            if v != 0:
                cartas_cheias.append(v)
            else:
                cartas_vazias.append(k)

        for carta in cartas_cheias:
            cartas_baralho_restante.remove(carta)

        for x in range(1, num_oponentes+1):
            cartas_vazias.append('carta1_adversario'+str(x))
            cartas_vazias.append('carta2_adversario'+str(x))

        for item in cartas_vazias:
            x = len(cartas_baralho_restante)
            num_randomico = random.randint(0, x-1)
            dicionario[item] = cartas_baralho_restante[num_randomico]
            cartas_baralho_restante.remove(cartas_baralho_restante[num_randomico])
        
        board = [dicionario['carta_mesa1'], dicionario['carta_mesa2'], dicionario['carta_mesa3'], dicionario['carta_mesa4'], dicionario['carta_mesa5']]
        
        hero_hand = list()
        for k,v in dicionario.items():
            if k not in ('carta_mesa1','carta_mesa2','carta_mesa3','carta_mesa4','carta_mesa5'):
                hero_hand.append(v)

        odds = holdem_calc.calculate(board, True, 1, None , hero_hand, False)

        soma_ganhar = soma_ganhar + odds[1]
        soma_empatar = soma_empatar + odds[0]

        i = i + 1

    soma_ganhar = soma_ganhar/n
    soma_empatar = soma_empatar/n
    if  num_oponentes == 1:
        pass
    elif num_oponentes == 2:
        soma_empatar = soma_empatar * 0.67
    elif num_oponentes == 3:
        soma_empatar = soma_empatar * 0.5
    elif num_oponentes == 4:
        soma_empatar = soma_empatar * 0.4
    elif num_oponentes == 5:
        soma_empatar = soma_empatar * 0.33
    elif num_oponentes == 6:
        soma_empatar = soma_empatar * 0.29
    elif num_oponentes == 7:
        soma_empatar = soma_empatar * 0.25
    elif num_oponentes == 8:
        soma_empatar = soma_empatar * 0.22
    elif num_oponentes == 9:
        soma_empatar = soma_empatar * 0.20
    
    return(soma_ganhar, soma_empatar)


def reconhecer_numero(image, lista_coordenadas):
    image_cortada = image.crop((lista_coordenadas[0] , lista_coordenadas[1], lista_coordenadas[2], lista_coordenadas[3])).convert('L')
    ret,img = cv2.threshold(np.array(image_cortada), 125, 255, cv2.THRESH_BINARY)
    img = Image.fromarray(img.astype(np.uint8))
    pot_text = pytesseract.image_to_string(img)
    x = re.findall('[0-9]+', pot_text)
    string = str()
    for numeros in x:
        string = string + numeros
    if string == '':
        x = 0
    else:
        x = int(string)
    return (x)


# (left=283, top=511, width=192, height=11) -- oponente posicao 2
# (left=332, top=182, width=192, height=11) -- oponente posicao 3
# (left=864, top=79, width=192, height=11) -- oponente posicao 4
# (left=1396, top=182, width=192, height=11) -- oponente posicao 5
# (left=1445, top=511, width=192, height=11) -- oponente posicao 6

def reconhecer_numero_oponentes(imagem):
    lista =list(pyautogui.locateAll(image_banco_carta_adversario, imagem, grayscale=False, confidence = 0.95))
    lista_posicoes = list()
    numero_jogadores = 0
    for x in lista:
        # print ('o valor de x eh: ', x)
        # print('o valor de x0 eh: ', x[0])
        if x[0] == 283:
            lista_posicoes.append(2)
            numero_jogadores = numero_jogadores + 1
        elif x[0] == 332:
            lista_posicoes.append(3)
            numero_jogadores = numero_jogadores + 1
        elif x[0] == 864:
            lista_posicoes.append(4)
            numero_jogadores = numero_jogadores + 1
        elif x[0] == 1396:
            lista_posicoes.append(5)
            numero_jogadores = numero_jogadores + 1
        elif x[0] == 1445:
            lista_posicoes.append(6)
            numero_jogadores = numero_jogadores + 1
        # elif x[0] == 1396:
        #     lista_posicoes.append(7)
        #     numero_jogadores = numero_jogadores + 1
        # elif x[0] == 1460:
        #     lista_posicoes.append(8)
        #     numero_jogadores = numero_jogadores + 1
        # elif x[0] == 1309:
        #     lista_posicoes.append(9)
        #     numero_jogadores = numero_jogadores + 1

    lista_posicoes.sort()
    
    return (lista_posicoes, numero_jogadores)


def reconhecer_numero_oponentes_na_mesa(imagem):
    lista =list(pyautogui.locateAll(image_oponentes_mesa, imagem, grayscale=False, confidence = 0.95))
    lista_posicoes = list()
    numero_jogadores = 0
    for x in lista:
        # print ('o valor de x eh: ', x)
        # print('o valor de x0 eh: ', x[0])
        if x[0] == 447:
            lista_posicoes.append(2)
            numero_jogadores = numero_jogadores + 1
        elif x[0] == 495:
            lista_posicoes.append(3)
            numero_jogadores = numero_jogadores + 1
        elif x[0] == 1028:
            lista_posicoes.append(4)
            numero_jogadores = numero_jogadores + 1
        elif x[0] == 1360:
            lista_posicoes.append(5)
            numero_jogadores = numero_jogadores + 1
        elif x[0] == 1409:
            lista_posicoes.append(6)
            numero_jogadores = numero_jogadores + 1
        # elif x[0] == 1396:
        #     lista_posicoes.append(7)
        #     numero_jogadores = numero_jogadores + 1
        # elif x[0] == 1460:
        #     lista_posicoes.append(8)
        #     numero_jogadores = numero_jogadores + 1
        # elif x[0] == 1309:
        #     lista_posicoes.append(9)
        #     numero_jogadores = numero_jogadores + 1

    lista_posicoes.sort()
    
    return (lista_posicoes, numero_jogadores)


def reconhecer__oponentes_ausentes(imagem):
    lista =list(pyautogui.locateAll(image_oponente_ausente, imagem, grayscale=False, confidence = 0.95))
    lista_posicoes = list()
    numero_jogadores = 0
    for x in lista:
        # print ('o valor de x eh: ', x)
        # print('o valor de x0 eh: ', x[0])
        if x[0] == 289:
            lista_posicoes.append(2)
            # numero_jogadores = numero_jogadores + 1
        elif x[0] == 337:
            lista_posicoes.append(3)
            # numero_jogadores = numero_jogadores + 1
        elif x[0] == 869:
            lista_posicoes.append(4)
            # numero_jogadores = numero_jogadores + 1
        elif x[0] == 1475:
            lista_posicoes.append(5)
            # numero_jogadores = numero_jogadores + 1
        elif x[0] == 1523:
            lista_posicoes.append(6)
            # numero_jogadores = numero_jogadores + 1
        # elif x[0] == 1396:
        #     lista_posicoes.append(7)
        #     numero_jogadores = numero_jogadores + 1
        # elif x[0] == 1460:
        #     lista_posicoes.append(8)
        #     numero_jogadores = numero_jogadores + 1
        # elif x[0] == 1309:
        #     lista_posicoes.append(9)
        #     numero_jogadores = numero_jogadores + 1

    lista_posicoes = list(dict.fromkeys(lista_posicoes))
    lista_posicoes.sort()
    numero_jogadores = len(lista_posicoes)
    
    return (lista_posicoes, numero_jogadores)


dicionario_classificacao_cartas = {
##-------------------------------------------------------------------------------------------------------------------
    ## offsuited!!
#AKo 5 (arrumado)
4.9: [['Ac', 'Kd'] , ['Ac', 'Kh'] , ['Ac', 'Ks'] , ['Ad', 'Kc'] , ['Ad', 'Kh'] , ['Ad', 'Ks'] , ['Ah', 'Kc'] , ['Ah', 'Kd'] , ['Ah', 'Ks'] , ['As', 'Kc'] , ['As', 'Kd'] , ['As', 'Kh']],
#AQo 8
8:[['Ac', 'Qd'] , ['Ac', 'Qh'] , ['Ac', 'Qs'] , ['Ad', 'Qc'] , ['Ad', 'Qh'] , ['Ad', 'Qs'] , ['Ah', 'Qc'] , ['Ah', 'Qd'] , ['Ah', 'Qs'] , ['As', 'Qc'] , ['As', 'Qd'] , ['As', 'Qh']],
#AJo 12
12:[['Ac', 'Jd'] , ['Ac', 'Jh'] , ['Ac', 'Js'] , ['Ad', 'Jc'] , ['Ad', 'Jh'] , ['Ad', 'Js'] , ['Ah', 'Jc'] , ['Ah', 'Jd'] , ['Ah', 'Js'] , ['As', 'Jc'] , ['As', 'Jd'] , ['As', 'Jh']],
#ATo 18
18:[['Ac', 'Td'] , ['Ac', 'Th'] , ['Ac', 'Ts'] , ['Ad', 'Tc'] , ['Ad', 'Th'] , ['Ad', 'Ts'] , ['Ah', 'Tc'] , ['Ah', 'Td'] , ['Ah', 'Ts'] , ['As', 'Tc'] , ['As', 'Td'] , ['As', 'Th']],
#A9o 32
32:[['Ac', '9d'] , ['Ac', '9h'] , ['Ac', '9s'] , ['Ad', '9c'] , ['Ad', '9h'] , ['Ad', '9s'] , ['Ah', '9c'] , ['Ah', '9d'] , ['Ah', '9s'] , ['As', '9c'] , ['As', '9d'] , ['As', '9h']],
#A8o 39 (arrumado)
38.9:[['Ac', '8d'] , ['Ac', '8h'] , ['Ac', '8s'] , ['Ad', '8c'] , ['Ad', '8h'] , ['Ad', '8s'] , ['Ah', '8c'] , ['Ah', '8d'] , ['Ah', '8s'] , ['As', '8c'] , ['As', '8d'] , ['As', '8h']],
#A7o 45
45:[['Ac', '7d'] , ['Ac', '7h'] , ['Ac', '7s'] , ['Ad', '7c'] , ['Ad', '7h'] , ['Ad', '7s'] , ['Ah', '7c'] , ['Ah', '7d'] , ['Ah', '7s'] , ['As', '7c'] , ['As', '7d'] , ['As', '7h']],
#A6o 51
51:[['Ac', '6d'] , ['Ac', '6h'] , ['Ac', '6s'] , ['Ad', '6c'] , ['Ad', '6h'] , ['Ad', '6s'] , ['Ah', '6c'] , ['Ah', '6d'] , ['Ah', '6s'] , ['As', '6c'] , ['As', '6d'] , ['As', '6h']],
#A5o 44
44:[['Ac', '5d'] , ['Ac', '5h'] , ['Ac', '5s'] , ['Ad', '5c'] , ['Ad', '5h'] , ['Ad', '5s'] , ['Ah', '5c'] , ['Ah', '5d'] , ['Ah', '5s'] , ['As', '5c'] , ['As', '5d'] , ['As', '5h']],
#A4o 46
46:[['Ac', '4d'] , ['Ac', '4h'] , ['Ac', '4s'] , ['Ad', '4c'] , ['Ad', '4h'] , ['Ad', '4s'] , ['Ah', '4c'] , ['Ah', '4d'] , ['Ah', '4s'] , ['As', '4c'] , ['As', '4d'] , ['As', '4h']],
#A3o 49
39:[['Ac', '3d'] , ['Ac', '3h'] , ['Ac', '3s'] , ['Ad', '3c'] , ['Ad', '3h'] , ['Ad', '3s'] , ['Ah', '3c'] , ['Ah', '3d'] , ['Ah', '3s'] , ['As', '3c'] , ['As', '3d'] , ['As', '3h']],
#A2o 54
54:[['Ac', '2d'] , ['Ac', '2h'] , ['Ac', '2s'] , ['Ad', '2c'] , ['Ad', '2h'] , ['Ad', '2s'] , ['Ah', '2c'] , ['Ah', '2d'] , ['Ah', '2s'] , ['As', '2c'] , ['As', '2d'] , ['As', '2h']],

#KQo 9
9:[['Kc', 'Qd'] , ['Kc', 'Qh'] , ['Kc', 'Qs'] , ['Kd', 'Qc'] , ['Kd', 'Qh'] , ['Kd', 'Qs'] , ['Kh', 'Qc'] , ['Kh', 'Qd'] , ['Kh', 'Qs'] , ['Ks', 'Qc'] , ['Ks', 'Qd'] , ['Ks', 'Qh']],
#KJo 14
14:[['Kc', 'Jd'] , ['Kc', 'Jh'] , ['Kc', 'Js'] , ['Kd', 'Jc'] , ['Kd', 'Jh'] , ['Kd', 'Js'] , ['Kh', 'Jc'] , ['Kh', 'Jd'] , ['Kh', 'Js'] , ['Ks', 'Jc'] , ['Ks', 'Jd'] , ['Ks', 'Jh']],
#KTo 20
20:[['Kc', 'Td'] , ['Kc', 'Th'] , ['Kc', 'Ts'] , ['Kd', 'Tc'] , ['Kd', 'Th'] , ['Kd', 'Ts'] , ['Kh', 'Tc'] , ['Kh', 'Td'] , ['Kh', 'Ts'] , ['Ks', 'Tc'] , ['Ks', 'Td'] , ['Ks', 'Th']],
#K9o 35
35:[['Kc', '9d'] , ['Kc', '9h'] , ['Kc', '9s'] , ['Kd', '9c'] , ['Kd', '9h'] , ['Kd', '9s'] , ['Kh', '9c'] , ['Kh', '9d'] , ['Kh', '9s'] , ['Ks', '9c'] , ['Ks', '9d'] , ['Ks', '9h']],
#K8o 50
50:[['Kc', '8d'] , ['Kc', '8h'] , ['Kc', '8s'] , ['Kd', '8c'] , ['Kd', '8h'] , ['Kd', '8s'] , ['Kh', '8c'] , ['Kh', '8d'] , ['Kh', '8s'] , ['Ks', '8c'] , ['Ks', '8d'] , ['Ks', '8h']],
#K7o 57 (arrumado)
56.9:[['Kc', '7d'] , ['Kc', '7h'] , ['Kc', '7s'] , ['Kd', '7c'] , ['Kd', '7h'] , ['Kd', '7s'] , ['Kh', '7c'] , ['Kh', '7d'] , ['Kh', '7s'] , ['Ks', '7c'] , ['Ks', '7d'] , ['Ks', '7h']],
#K6o 60
60:[['Kc', '6d'] , ['Kc', '6h'] , ['Kc', '6s'] , ['Kd', '6c'] , ['Kd', '6h'] , ['Kd', '6s'] , ['Kh', '6c'] , ['Kh', '6d'] , ['Kh', '6s'] , ['Ks', '6c'] , ['Ks', '6d'] , ['Ks', '6h']],
#K5o 63
63:[['Kc', '5d'] , ['Kc', '5h'] , ['Kc', '5s'] , ['Kd', '5c'] , ['Kd', '5h'] , ['Kd', '5s'] , ['Kh', '5c'] , ['Kh', '5d'] , ['Kh', '5s'] , ['Ks', '5c'] , ['Ks', '5d'] , ['Ks', '5h']],
#K4o 67 (arrumado)
66.9:[['Kc', '4d'] , ['Kc', '4h'] , ['Kc', '4s'] , ['Kd', '4c'] , ['Kd', '4h'] , ['Kd', '4s'] , ['Kh', '4c'] , ['Kh', '4d'] , ['Kh', '4s'] , ['Ks', '4c'] , ['Ks', '4d'] , ['Ks', '4h']],
#K3o 67
67:[['Kc', '3d'] , ['Kc', '3h'] , ['Kc', '3s'] , ['Kd', '3c'] , ['Kd', '3h'] , ['Kd', '3s'] , ['Kh', '3c'] , ['Kh', '3d'] , ['Kh', '3s'] , ['Ks', '3c'] , ['Ks', '3d'] , ['Ks', '3h']],
#K2o 69
69:[['Kc', '2d'] , ['Kc', '2h'] , ['Kc', '2s'] , ['Kd', '2c'] , ['Kd', '2h'] , ['Kd', '2s'] , ['Kh', '2c'] , ['Kh', '2d'] , ['Kh', '2s'] , ['Ks', '2c'] , ['Ks', '2d'] , ['Ks', '2h']],

#QJo 15
15:[['Qc', 'Jd'] , ['Qc', 'Jh'] , ['Qc', 'Js'] , ['Qd', 'Jc'] , ['Qd', 'Jh'] , ['Qd', 'Js'] , ['Qh', 'Jc'] , ['Qh', 'Jd'] , ['Qh', 'Js'] , ['Qs', 'Jc'] , ['Qs', 'Jd'] , ['Qs', 'Jh']],
#QTo 22
22:[['Qc', 'Td'] , ['Qc', 'Th'] , ['Qc', 'Ts'] , ['Qd', 'Tc'] , ['Qd', 'Th'] , ['Qd', 'Ts'] , ['Qh', 'Tc'] , ['Qh', 'Td'] , ['Qh', 'Ts'] , ['Qs', 'Tc'] , ['Qs', 'Td'] , ['Qs', 'Th']],
#Q9o 36
36:[['Qc', '9d'] , ['Qc', '9h'] , ['Qc', '9s'] , ['Qd', '9c'] , ['Qd', '9h'] , ['Qd', '9s'] , ['Qh', '9c'] , ['Qh', '9d'] , ['Qh', '9s'] , ['Qs', '9c'] , ['Qs', '9d'] , ['Qs', '9h']],
#Q8o 53
53:[['Qc', '8d'] , ['Qc', '8h'] , ['Qc', '8s'] , ['Qd', '8c'] , ['Qd', '8h'] , ['Qd', '8s'] , ['Qh', '8c'] , ['Qh', '8d'] , ['Qh', '8s'] , ['Qs', '8c'] , ['Qs', '8d'] , ['Qs', '8h']],
#Q7o 66
66:[['Qc', '7d'] , ['Qc', '7h'] , ['Qc', '7s'] , ['Qd', '7c'] , ['Qd', '7h'] , ['Qd', '7s'] , ['Qh', '7c'] , ['Qh', '7d'] , ['Qh', '7s'] , ['Qs', '7c'] , ['Qs', '7d'] , ['Qs', '7h']],
#Q6o 71
71:[['Qc', '6d'] , ['Qc', '6h'] , ['Qc', '6s'] , ['Qd', '6c'] , ['Qd', '6h'] , ['Qd', '6s'] , ['Qh', '6c'] , ['Qh', '6d'] , ['Qh', '6s'] , ['Qs', '6c'] , ['Qs', '6d'] , ['Qs', '6h']],
#Q5o 75
75:[['Qc', '5d'] , ['Qc', '5h'] , ['Qc', '5s'] , ['Qd', '5c'] , ['Qd', '5h'] , ['Qd', '5s'] , ['Qh', '5c'] , ['Qh', '5d'] , ['Qh', '5s'] , ['Qs', '5c'] , ['Qs', '5d'] , ['Qs', '5h']],
#Q4o 76 (arrumado)
75.9:[['Qc', '4d'] , ['Qc', '4h'] , ['Qc', '4s'] , ['Qd', '4c'] , ['Qd', '4h'] , ['Qd', '4s'] , ['Qh', '4c'] , ['Qh', '4d'] , ['Qh', '4s'] , ['Qs', '4c'] , ['Qs', '4d'] , ['Qs', '4h']],
#Q3o 77
77:[['Qc', '3d'] , ['Qc', '3h'] , ['Qc', '3s'] , ['Qd', '3c'] , ['Qd', '3h'] , ['Qd', '3s'] , ['Qh', '3c'] , ['Qh', '3d'] , ['Qh', '3s'] , ['Qs', '3c'] , ['Qs', '3d'] , ['Qs', '3h']],
#Q2o 79
79:[['Qc', '2d'] , ['Qc', '2h'] , ['Qc', '2s'] , ['Qd', '2c'] , ['Qd', '2h'] , ['Qd', '2s'] , ['Qh', '2c'] , ['Qh', '2d'] , ['Qh', '2s'] , ['Qs', '2c'] , ['Qs', '2d'] , ['Qs', '2h']],

#JTo 21
21:[['Jc', 'Td'] , ['Jc', 'Th'] , ['Jc', 'Ts'] , ['Jd', 'Tc'] , ['Jd', 'Th'] , ['Jd', 'Ts'] , ['Jh', 'Tc'] , ['Jh', 'Td'] , ['Jh', 'Ts'] , ['Js', 'Tc'] , ['Js', 'Td'] , ['Js', 'Th']],
#J9o 34 (arrumado)
33.9:[['Jc', '9d'] , ['Jc', '9h'] , ['Jc', '9s'] , ['Jd', '9c'] , ['Jd', '9h'] , ['Jd', '9s'] , ['Jh', '9c'] , ['Jh', '9d'] , ['Jh', '9s'] , ['Js', '9c'] , ['Js', '9d'] , ['Js', '9h']],
#J8o 48
48:[['Jc', '8d'] , ['Jc', '8h'] , ['Jc', '8s'] , ['Jd', '8c'] , ['Jd', '8h'] , ['Jd', '8s'] , ['Jh', '8c'] , ['Jh', '8d'] , ['Jh', '8s'] , ['Js', '8c'] , ['Js', '8d'] , ['Js', '8h']],
#J7o 64
64:[['Jc', '7d'] , ['Jc', '7h'] , ['Jc', '7s'] , ['Jd', '7c'] , ['Jd', '7h'] , ['Jd', '7s'] , ['Jh', '7c'] , ['Jh', '7d'] , ['Jh', '7s'] , ['Js', '7c'] , ['Js', '7d'] , ['Js', '7h']],
#J6o 80
80:[['Jc', '6d'] , ['Jc', '6h'] , ['Jc', '6s'] , ['Jd', '6c'] , ['Jd', '6h'] , ['Jd', '6s'] , ['Jh', '6c'] , ['Jh', '6d'] , ['Jh', '6s'] , ['Js', '6c'] , ['Js', '6d'] , ['Js', '6h']],
#J5o 82
82:[['Jc', '5d'] , ['Jc', '5h'] , ['Jc', '5s'] , ['Jd', '5c'] , ['Jd', '5h'] , ['Jd', '5s'] , ['Jh', '5c'] , ['Jh', '5d'] , ['Jh', '5s'] , ['Js', '5c'] , ['Js', '5d'] , ['Js', '5h']],
#J4o 85
85:[['Jc', '4d'] , ['Jc', '4h'] , ['Jc', '4s'] , ['Jd', '4c'] , ['Jd', '4h'] , ['Jd', '4s'] , ['Jh', '4c'] , ['Jh', '4d'] , ['Jh', '4s'] , ['Js', '4c'] , ['Js', '4d'] , ['Js', '4h']],
#J3o 86 (arrumado)
85.9:[['Jc', '3d'] , ['Jc', '3h'] , ['Jc', '3s'] , ['Jd', '3c'] , ['Jd', '3h'] , ['Jd', '3s'] , ['Jh', '3c'] , ['Jh', '3d'] , ['Jh', '3s'] , ['Js', '3c'] , ['Js', '3d'] , ['Js', '3h']],
#J2o 87
87:[['Jc', '2d'] , ['Jc', '2h'] , ['Jc', '2s'] , ['Jd', '2c'] , ['Jd', '2h'] , ['Jd', '2s'] , ['Jh', '2c'] , ['Jh', '2d'] , ['Jh', '2s'] , ['Js', '2c'] , ['Js', '2d'] , ['Js', '2h']],

#T9o 31
31:[['Tc', '9d'] , ['Tc', '9h'] , ['Tc', '9s'] , ['Td', '9c'] , ['Td', '9h'] , ['Td', '9s'] , ['Th', '9c'] , ['Th', '9d'] , ['Th', '9s'] , ['Ts', '9c'] , ['Ts', '9d'] , ['Ts', '9h']],
#T8o 43
43:[['Tc', '8d'] , ['Tc', '8h'] , ['Tc', '8s'] , ['Td', '8c'] , ['Td', '8h'] , ['Td', '8s'] , ['Th', '8c'] , ['Th', '8d'] , ['Th', '8s'] , ['Ts', '8c'] , ['Ts', '8d'] , ['Ts', '8h']],
#T7o 59
59:[['Tc', '7d'] , ['Tc', '7h'] , ['Tc', '7s'] , ['Td', '7c'] , ['Td', '7h'] , ['Td', '7s'] , ['Th', '7c'] , ['Th', '7d'] , ['Th', '7s'] , ['Ts', '7c'] , ['Ts', '7d'] , ['Ts', '7h']],
#T6o 74
74:[['Tc', '6d'] , ['Tc', '6h'] , ['Tc', '6s'] , ['Td', '6c'] , ['Td', '6h'] , ['Td', '6s'] , ['Th', '6c'] , ['Th', '6d'] , ['Th', '6s'] , ['Ts', '6c'] , ['Ts', '6d'] , ['Ts', '6h']],
#T5o 89
89:[['Tc', '5d'] , ['Tc', '5h'] , ['Tc', '5s'] , ['Td', '5c'] , ['Td', '5h'] , ['Td', '5s'] , ['Th', '5c'] , ['Th', '5d'] , ['Th', '5s'] , ['Ts', '5c'] , ['Ts', '5d'] , ['Ts', '5h']],
#T4o 90
90:[['Tc', '4d'] , ['Tc', '4h'] , ['Tc', '4s'] , ['Td', '4c'] , ['Td', '4h'] , ['Td', '4s'] , ['Th', '4c'] , ['Th', '4d'] , ['Th', '4s'] , ['Ts', '4c'] , ['Ts', '4d'] , ['Ts', '4h']],
#T3o 92
92:[['Tc', '3d'] , ['Tc', '3h'] , ['Tc', '3s'] , ['Td', '3c'] , ['Td', '3h'] , ['Td', '3s'] , ['Th', '3c'] , ['Th', '3d'] , ['Th', '3s'] , ['Ts', '3c'] , ['Ts', '3d'] , ['Ts', '3h']],
#T2o 94
94:[['Tc', '2d'] , ['Tc', '2h'] , ['Tc', '2s'] , ['Td', '2c'] , ['Td', '2h'] , ['Td', '2s'] , ['Th', '2c'] , ['Th', '2d'] , ['Th', '2s'] , ['Ts', '2c'] , ['Ts', '2d'] , ['Ts', '2h']],

#98o 42
42:[['9c', '8d'] , ['9c', '8h'] , ['9c', '8s'] , ['9d', '8c'] , ['9d', '8h'] , ['9d', '8s'] , ['9h', '8c'] , ['9h', '8d'] , ['9h', '8s'] , ['9s', '8c'] , ['9s', '8d'] , ['9s', '8h']],
#97o 55
55:[['9c', '7d'] , ['9c', '7h'] , ['9c', '7s'] , ['9d', '7c'] , ['9d', '7h'] , ['9d', '7s'] , ['9h', '7c'] , ['9h', '7d'] , ['9h', '7s'] , ['9s', '7c'] , ['9s', '7d'] , ['9s', '7h']],
#96o 68
68:[['9c', '6d'] , ['9c', '6h'] , ['9c', '6s'] , ['9d', '6c'] , ['9d', '6h'] , ['9d', '6s'] , ['9h', '6c'] , ['9h', '6d'] , ['9h', '6s'] , ['9s', '6c'] , ['9s', '6d'] , ['9s', '6h']],
#95o 83
83:[['9c', '5d'] , ['9c', '5h'] , ['9c', '5s'] , ['9d', '5c'] , ['9d', '5h'] , ['9d', '5s'] , ['9h', '5c'] , ['9h', '5d'] , ['9h', '5s'] , ['9s', '5c'] , ['9s', '5d'] , ['9s', '5h']],
#94o 95 (arrumado)
94.9:[['9c', '4d'] , ['9c', '4h'] , ['9c', '4s'] , ['9d', '4c'] , ['9d', '4h'] , ['9d', '4s'] , ['9h', '4c'] , ['9h', '4d'] , ['9h', '4s'] , ['9s', '4c'] , ['9s', '4d'] , ['9s', '4h']],
#93o 96
96:[['9c', '3d'] , ['9c', '3h'] , ['9c', '3s'] , ['9d', '3c'] , ['9d', '3h'] , ['9d', '3s'] , ['9h', '3c'] , ['9h', '3d'] , ['9h', '3s'] , ['9s', '3c'] , ['9s', '3d'] , ['9s', '3h']],
#92o 97
97:[['9c', '2d'] , ['9c', '2h'] , ['9c', '2s'] , ['9d', '2c'] , ['9d', '2h'] , ['9d', '2s'] , ['9h', '2c'] , ['9h', '2d'] , ['9h', '2s'] , ['9s', '2c'] , ['9s', '2d'] , ['9s', '2h']],

#87o 52
52:[['8c', '7d'] , ['8c', '7h'] , ['8c', '7s'] , ['8d', '7c'] , ['8d', '7h'] , ['8d', '7s'] , ['8h', '7c'] , ['8h', '7d'] , ['8h', '7s'] , ['8s', '7c'] , ['8s', '7d'] , ['8s', '7h']],
#86o 61
61:[['8c', '6d'] , ['8c', '6h'] , ['8c', '6s'] , ['8d', '6c'] , ['8d', '6h'] , ['8d', '6s'] , ['8h', '6c'] , ['8h', '6d'] , ['8h', '6s'] , ['8s', '6c'] , ['8s', '6d'] , ['8s', '6h']],
#85o 73
73:[['8c', '5d'] , ['8c', '5h'] , ['8c', '5s'] , ['8d', '5c'] , ['8d', '5h'] , ['8d', '5s'] , ['8h', '5c'] , ['8h', '5d'] , ['8h', '5s'] , ['8s', '5c'] , ['8s', '5d'] , ['8s', '5h']],
#84o 88
88:[['8c', '4d'] , ['8c', '4h'] , ['8c', '4s'] , ['8d', '4c'] , ['8d', '4h'] , ['8d', '4s'] , ['8h', '4c'] , ['8h', '4d'] , ['8h', '4s'] , ['8s', '4c'] , ['8s', '4d'] , ['8s', '4h']],
#83o 98
98:[['8c', '3d'] , ['8c', '3h'] , ['8c', '3s'] , ['8d', '3c'] , ['8d', '3h'] , ['8d', '3s'] , ['8h', '3c'] , ['8h', '3d'] , ['8h', '3s'] , ['8s', '3c'] , ['8s', '3d'] , ['8s', '3h']],
#82o 99
99:[['8c', '2d'] , ['8c', '2h'] , ['8c', '2s'] , ['8d', '2c'] , ['8d', '2h'] , ['8d', '2s'] , ['8h', '2c'] , ['8h', '2d'] , ['8h', '2s'] , ['8s', '2c'] , ['8s', '2d'] , ['8s', '2h']],

#76o 57
57:[['7c', '6d'] , ['7c', '6h'] , ['7c', '6s'] , ['7d', '6c'] , ['7d', '6h'] , ['7d', '6s'] , ['7h', '6c'] , ['7h', '6d'] , ['7h', '6s'] , ['7s', '6c'] , ['7s', '6d'] , ['7s', '6h']],
#75o 65
65:[['7c', '5d'] , ['7c', '5h'] , ['7c', '5s'] , ['7d', '5c'] , ['7d', '5h'] , ['7d', '5s'] , ['7h', '5c'] , ['7h', '5d'] , ['7h', '5s'] , ['7s', '5c'] , ['7s', '5d'] , ['7s', '5h']],
#74o 78
78:[['7c', '4d'] , ['7c', '4h'] , ['7c', '4s'] , ['7d', '4c'] , ['7d', '4h'] , ['7d', '4s'] , ['7h', '4c'] , ['7h', '4d'] , ['7h', '4s'] , ['7s', '4c'] , ['7s', '4d'] , ['7s', '4h']],
#73o 93
93:[['7c', '3d'] , ['7c', '3h'] , ['7c', '3s'] , ['7d', '3c'] , ['7d', '3h'] , ['7d', '3s'] , ['7h', '3c'] , ['7h', '3d'] , ['7h', '3s'] , ['7s', '3c'] , ['7s', '3d'] , ['7s', '3h']],
#72o 100
100:[['7c', '2d'] , ['7c', '2h'] , ['7c', '2s'] , ['7d', '2c'] , ['7d', '2h'] , ['7d', '2s'] , ['7h', '2c'] , ['7h', '2d'] , ['7h', '2s'] , ['7s', '2c'] , ['7s', '2d'] , ['7s', '2h']],

#65o 58
58:[['6c', '5d'] , ['6c', '5h'] , ['6c', '5s'] , ['6d', '5c'] , ['6d', '5h'] , ['6d', '5s'] , ['6h', '5c'] , ['6h', '5d'] , ['6h', '5s'] , ['6s', '5c'] , ['6s', '5d'] , ['6s', '5h']],
#64o 70
70:[['6c', '4d'] , ['6c', '4h'] , ['6c', '4s'] , ['6d', '4c'] , ['6d', '4h'] , ['6d', '4s'] , ['6h', '4c'] , ['6h', '4d'] , ['6h', '4s'] , ['6s', '4c'] , ['6s', '4d'] , ['6s', '4h']],
#63o 81
81:[['6c', '3d'] , ['6c', '3h'] , ['6c', '3s'] , ['6d', '3c'] , ['6d', '3h'] , ['6d', '3s'] , ['6h', '3c'] , ['6h', '3d'] , ['6h', '3s'] , ['6s', '3c'] , ['6s', '3d'] , ['6s', '3h']],
#62o 95
95:[['6c', '2d'] , ['6c', '2h'] , ['6c', '2s'] , ['6d', '2c'] , ['6d', '2h'] , ['6d', '2s'] , ['6h', '2c'] , ['6h', '2d'] , ['6h', '2s'] , ['6s', '2c'] , ['6s', '2d'] , ['6s', '2h']],

#54o 62
62:[['5c', '4d'] , ['5c', '4h'] , ['5c', '4s'] , ['5d', '4c'] , ['5d', '4h'] , ['5d', '4s'] , ['5h', '4c'] , ['5h', '4d'] , ['5h', '4s'] , ['5s', '4c'] , ['5s', '4d'] , ['5s', '4h']],
#53o 72
72:[['5c', '3d'] , ['5c', '3h'] , ['5c', '3s'] , ['5d', '3c'] , ['5d', '3h'] , ['5d', '3s'] , ['5h', '3c'] , ['5h', '3d'] , ['5h', '3s'] , ['5s', '3c'] , ['5s', '3d'] , ['5s', '3h']],
#52o 84
84:[['5c', '2d'] , ['5c', '2h'] , ['5c', '2s'] , ['5d', '2c'] , ['5d', '2h'] , ['5d', '2s'] , ['5h', '2c'] , ['5h', '2d'] , ['5h', '2s'] , ['5s', '2c'] , ['5s', '2d'] , ['5s', '2h']],

#43o 76
76:[['4c', '3d'] , ['4c', '3h'] , ['4c', '3s'] , ['4d', '3c'] , ['4d', '3h'] , ['4d', '3s'] , ['4h', '3c'] , ['4h', '3d'] , ['4h', '3s'] , ['4s', '3c'] , ['4s', '3d'] , ['4s', '3h']],
#42o 86
86:[['4c', '2d'] , ['4c', '2h'] , ['4c', '2s'] , ['4d', '2c'] , ['4d', '2h'] , ['4d', '2s'] , ['4h', '2c'] , ['4h', '2d'] , ['4h', '2s'] , ['4s', '2c'] , ['4s', '2d'] , ['4s', '2h']],

#32o 91
91:[['3c', '2d'] , ['3c', '2h'] , ['3c', '2s'] , ['3d', '2c'] , ['3d', '2h'] , ['3d', '2s'] , ['3h', '2c'] , ['3h', '2d'] , ['3h', '2s'] , ['3s', '2c'] , ['3s', '2d'] , ['3s', '2h']],


##-------------------------------------------------------------------------------------------------------------------
## suited!!
#AKs 2
1.9:[['Ac', 'Kc'] , ['Ad', 'Kd'] , ['Ah', 'Kh'] , ['As', 'Ks']],
#AQs 2
2:[['Ac', 'Qc'] , ['Ad', 'Qd'] , ['Ah', 'Qh'] , ['As', 'Qs']],
#AJs 3 (arrumado)
2.9:[['Ac', 'Jc'] , ['Ad', 'Jd'] , ['Ah', 'Jh'] , ['As', 'Js']],
#ATs 5 (arrumado)
4.8:[['Ac', 'Tc'] , ['Ad', 'Td'] , ['Ah', 'Th'] , ['As', 'Ts']],
#A9s 8 (arrumado)
7.9:[['Ac', '9c'] , ['Ad', '9d'] , ['Ah', '9h'] , ['As', '9s']],
#A8s 10 (arrumado)
9.7:[['Ac', '8c'] , ['Ad', '8d'] , ['Ah', '8h'] , ['As', '8s']],
#A7s 13
13:[['Ac', '7c'] , ['Ad', '7d'] , ['Ah', '7h'] , ['As', '7s']],
#A6s 14 (arrumado)
13.7:[['Ac', '6c'] , ['Ad', '6d'] , ['Ah', '6h'] , ['As', '6s']],
#A5s 12 (arrumado)
11.9:[['Ac', '5c'] , ['Ad', '5d'] , ['Ah', '5h'] , ['As', '5s']],
#A4s 14 (arrumado)
13.9:[['Ac', '4c'] , ['Ad', '4d'] , ['Ah', '4h'] , ['As', '4s']],
#A3s 14 (arrumado)
13.8:[['Ac', '3c'] , ['Ad', '3d'] , ['Ah', '3h'] , ['As', '3s']],
#A2s 17 (arrumado)
16.9:[['Ac', '2c'] , ['Ad', '2d'] , ['Ah', '2h'] , ['As', '2s']],

#KQs 3
2.8:[['Kc', 'Qc'] , ['Kd', 'Qd'] , ['Kh', 'Qh'] , ['Ks', 'Qs']],
#KJs 3
3:[['Kc', 'Jc'] , ['Kd', 'Jd'] , ['Kh', 'Jh'] , ['Ks', 'Js']],
#KTs 6 (arrumado)
5.9:[['Kc', 'Tc'] , ['Kd', 'Td'] , ['Kh', 'Th'] , ['Ks', 'Ts']],
#K9s 10 (arrumado)
9.9:[['Kc', '9c'] , ['Kd', '9d'] , ['Kh', '9h'] , ['Ks', '9s']],
#K8s 16 (arrumado)
15.9:[['Kc', '8c'] , ['Kd', '8d'] , ['Kh', '8h'] , ['Ks', '8s']],
#K7s 19 (arrumado)
18.9:[['Kc', '7c'] , ['Kd', '7d'] , ['Kh', '7h'] , ['Ks', '7s']],
#K6s 24 (arrumado)
23.9:[['Kc', '6c'] , ['Kd', '6d'] , ['Kh', '6h'] , ['Ks', '6s']],
#K5s 25 (arrumado)
24.9:[['Kc', '5c'] , ['Kd', '5d'] , ['Kh', '5h'] , ['Ks', '5s']],
#K4s 25 (arrumado)
24.8:[['Kc', '4c'] , ['Kd', '4d'] , ['Kh', '4h'] , ['Ks', '4s']],
#K3s 26 (arrumado)
25.9:[['Kc', '3c'] , ['Kd', '3d'] , ['Kh', '3h'] , ['Ks', '3s']],
#K2s 26 (arrumado)
25.8:[['Kc', '2c'] , ['Kd', '2d'] , ['Kh', '2h'] , ['Ks', '2s']],

#QJs 5
5:[['Qc', 'Jc'] , ['Qd', 'Jd'] , ['Qh', 'Jh'] , ['Qs', 'Js']],
#QTs 6 (arrumado)
5.8:[['Qc', 'Tc'] , ['Qd', 'Td'] , ['Qh', 'Th'] , ['Qs', 'Ts']],
#Q9s 10 (arrumado)
9.8:[['Qc', '9c'] , ['Qd', '9d'] , ['Qh', '9h'] , ['Qs', '9s']],
#Q8s 19
19:[['Qc', '8c'] , ['Qd', '8d'] , ['Qh', '8h'] , ['Qs', '8s']],
#Q7s 26
26:[['Qc', '7c'] , ['Qd', '7d'] , ['Qh', '7h'] , ['Qs', '7s']],
#Q6s 28 (arrumado)
27.9:[['Qc', '6c'] , ['Qd', '6d'] , ['Qh', '6h'] , ['Qs', '6s']],
#Q5s 29 (arrumado)
28.9:[['Qc', '5c'] , ['Qd', '5d'] , ['Qh', '5h'] , ['Qs', '5s']],
#Q4s 29 (arrumado)
28.8:[['Qc', '4c'] , ['Qd', '4d'] , ['Qh', '4h'] , ['Qs', '4s']],
#Q3s 30
30:[['Qc', '3c'] , ['Qd', '3d'] , ['Qh', '3h'] , ['Qs', '3s']],
#Q2s 31 (arrumado)
30.9:[['Qc', '2c'] , ['Qd', '2d'] , ['Qh', '2h'] , ['Qs', '2s']],

#JTs 6
6:[['Jc', 'Tc'] , ['Jd', 'Td'] , ['Jh', 'Th'] , ['Js', 'Ts']],
#J9s 11
11:[['Jc', '9c'] , ['Jd', '9d'] , ['Jh', '9h'] , ['Js', '9s']],
#J8s 17 (arrumado)
16.8:[['Jc', '8c'] , ['Jd', '8d'] , ['Jh', '8h'] , ['Js', '8s']],
#J7s 27 (arrumado)
26.9:[['Jc', '7c'] , ['Jd', '7d'] , ['Jh', '7h'] , ['Js', '7s']],
#J6s 33 (arrumado)
32.9:[['Jc', '6c'] , ['Jd', '6d'] , ['Jh', '6h'] , ['Js', '6s']],
#J5s 35 (arrumado)
34.9:[['Jc', '5c'] , ['Jd', '5d'] , ['Jh', '5h'] , ['Js', '5s']],
#J4s 37 (arrumado)
36.9:[['Jc', '4c'] , ['Jd', '4d'] , ['Jh', '4h'] , ['Js', '4s']],
#J3s 37 (arrumado)
36.8:[['Jc', '3c'] , ['Jd', '3d'] , ['Jh', '3h'] , ['Js', '3s']],
#J2s 38
37.9:[['Jc', '2c'] , ['Jd', '2d'] , ['Jh', '2h'] , ['Js', '2s']],

#T9s 10
10:[['Tc', '9c'] , ['Td', '9d'] , ['Th', '9h'] , ['Ts', '9s']],
#T8s 16 (arrumado)
15.8:[['Tc', '8c'] , ['Td', '8d'] , ['Th', '8h'] , ['Ts', '8s']],
#T7s 25 (arrumado)
24.8:[['Tc', '7c'] , ['Td', '7d'] , ['Th', '7h'] , ['Ts', '7s']],
#T6s 31 (arrumado)
30.8:[['Tc', '6c'] , ['Td', '6d'] , ['Th', '6h'] , ['Ts', '6s']],
#T5s 40 (arrumado)
39.9:[['Tc', '5c'] , ['Td', '5d'] , ['Th', '5h'] , ['Ts', '5s']],
#T4s 40 (arrumado)
39.8:[['Tc', '4c'] , ['Td', '4d'] , ['Th', '4h'] , ['Ts', '4s']],
#T3s 41
40.9:[['Tc', '3c'] , ['Td', '3d'] , ['Th', '3h'] , ['Ts', '3s']],
#T2s 41
40.8:[['Tc', '2c'] , ['Td', '2d'] , ['Th', '2h'] , ['Ts', '2s']],

#98s 17 
17:[['9c', '8c'] , ['9d', '8d'] , ['9h', '8h'] , ['9s', '8s']],
#97s 24 (arrumado)
23.8:[['9c', '7c'] , ['9d', '7d'] , ['9h', '7h'] , ['9s', '7s']],
#96s 29 (arrumado)
28.7:[['9c', '6c'] , ['9d', '6d'] , ['9h', '6h'] , ['9s', '6s']],
#95s 38
37.8:[['9c', '5c'] , ['9d', '5d'] , ['9h', '5h'] , ['9s', '5s']],
#94s 47 (arrumado)
46.9:[['9c', '4c'] , ['9d', '4d'] , ['9h', '4h'] , ['9s', '4s']],
#93s 47
47:[['9c', '3c'] , ['9d', '3d'] , ['9h', '3h'] , ['9s', '3s']],
#92s 49 (arrumado)
48.9:[['9c', '2c'] , ['9d', '2d'] , ['9h', '2h'] , ['9s', '2s']],

#87s 21 (arrumado)
20.9:[['8c', '7c'] , ['8d', '7d'] , ['8h', '7h'] , ['8s', '7s']],
#86s 27 (arrumado)
26.8:[['8c', '6c'] , ['8d', '6d'] , ['8h', '6h'] , ['8s', '6s']],
#85s 33
33:[['8c', '5c'] , ['8d', '5d'] , ['8h', '5h'] , ['8s', '5s']],
#84s 40
40:[['8c', '4c'] , ['8d', '4d'] , ['8h', '4h'] , ['8s', '4s']],
#83s 53 (arrumado)
52.9:[['8c', '3c'] , ['8d', '3d'] , ['8h', '3h'] , ['8s', '3s']],
#82s 54 (arrumado)
53.9:[['8c', '2c'] , ['8d', '2d'] , ['8h', '2h'] , ['8s', '2s']],

#76s 25
25:[['7c', '6c'] , ['7d', '6d'] , ['7h', '6h'] , ['7s', '6s']],
#75s 28 (arrumado)
27.8:[['7c', '5c'] , ['7d', '5d'] , ['7h', '5h'] , ['7s', '5s']],
#74s 37
37:[['7c', '4c'] , ['7d', '4d'] , ['7h', '4h'] , ['7s', '4s']],
#73s 45 (arrumado)
44.9:[['7c', '3c'] , ['7d', '3d'] , ['7h', '3h'] , ['7s', '3s']],
#72s 56
56:[['7c', '2c'] , ['7d', '2d'] , ['7h', '2h'] , ['7s', '2s']],

#65s 27
27:[['6c', '5c'] , ['6d', '5d'] , ['6h', '5h'] , ['6s', '5s']],
#64s 29
29:[['6c', '4c'] , ['6d', '4d'] , ['6h', '4h'] , ['6s', '4s']],
#63s 38
38:[['6c', '3c'] , ['6d', '3d'] , ['6h', '3h'] , ['6s', '3s']],
#62s 49 (arrumado)
48.8:[['6c', '2c'] , ['6d', '2d'] , ['6h', '2h'] , ['6s', '2s']],

#54s 28
28:[['5c', '4c'] , ['5d', '4d'] , ['5h', '4h'] , ['5s', '4s']],
#53s 32 (arrumado)
31.9:[['5c', '3c'] , ['5d', '3d'] , ['5h', '3h'] , ['5s', '3s']],
#52s 39 (arrumado)
38.9:[['5c', '2c'] , ['5d', '2d'] , ['5h', '2h'] , ['5s', '2s']],

#43s 36 (arrumado)
35.9:[['4c', '3c'] , ['4d', '3d'] , ['4h', '3h'] , ['4s', '3s']],
#42s 41
41:[['4c', '2c'] , ['4d', '2d'] , ['4h', '2h'] , ['4s', '2s']],

#32s 46 (arrumado)
45.9:[['3c', '2c'] , ['3d', '2d'] , ['3h', '2h'] , ['3s', '2s']],


##-----------------------------------------------------------------------------------------------------
## pares!!! esse tá certo!
#AA 0
0:[['Ac', 'Ad'] , ['Ac', 'Ah'] , ['Ac', 'As'] , ['Ad', 'Ah'] , ['Ad', 'As'] , ['Ah', 'As']],
#KK 1 (arrumado)
0.9:[['Kc', 'Kd'] , ['Kc', 'Kh'] , ['Kc', 'Ks'] , ['Kd', 'Kh'] , ['Kd', 'Ks'] , ['Kh', 'Ks']],
#QQ 1
1:[['Qc', 'Qd'] , ['Qc', 'Qh'] , ['Qc', 'Qs'] , ['Qd', 'Qh'] , ['Qd', 'Qs'] , ['Qh', 'Qs']],
#JJ 2 (arrumado)
1.8:[['Jc', 'Jd'] , ['Jc', 'Jh'] , ['Jc', 'Js'] , ['Jd', 'Jh'] , ['Jd', 'Js'] , ['Jh', 'Js']],
#TT 4
4:[['Tc', 'Td'] , ['Tc', 'Th'] , ['Tc', 'Ts'] , ['Td', 'Th'] , ['Td', 'Ts'] , ['Th', 'Ts']],
#99 7
7:[['9c', '9d'] , ['9c', '9h'] , ['9c', '9s'] , ['9d', '9h'] , ['9d', '9s'] , ['9h', '9s']],
#88 9 (arrumado)
8.9:[['8c', '8d'] , ['8c', '8h'] , ['8c', '8s'] , ['8d', '8h'] , ['8d', '8s'] , ['8h', '8s']],
#77 12 (arrumado)
11.8:[['7c', '7d'] , ['7c', '7h'] , ['7c', '7s'] , ['7d', '7h'] , ['7d', '7s'] , ['7h', '7s']],
#66 16
16:[['6c', '6d'] , ['6c', '6h'] , ['6c', '6s'] , ['6d', '6h'] , ['6d', '6s'] , ['6h', '6s']],
#55 20 (arrumado)
19.9:[['5c', '5d'] , ['5c', '5h'] , ['5c', '5s'] , ['5d', '5h'] , ['5d', '5s'] , ['5h', '5s']],
#44 23 (arrumado)
22.9:[['4c', '4d'] , ['4c', '4h'] , ['4c', '4s'] , ['4d', '4h'] , ['4d', '4s'] , ['4h', '4s']],
#33 23
23:[['3c', '3d'] , ['3c', '3h'] , ['3c', '3s'] , ['3d', '3h'] , ['3d', '3s'] , ['3h', '3s']],
#22 24
24:[['2c', '2d'] , ['2c', '2h'] , ['2c', '2s'] , ['2d', '2h'] , ['2d', '2s'] , ['2h', '2s']]
}

def classificar_mao(minha_mao):
    x = 101
    for k, v in dicionario_classificacao_cartas.items():
        if minha_mao in v:
            x = k
    
    if x == 101:
        minha_mao = [minha_mao[1], minha_mao[0]]
        for k, v in dicionario_classificacao_cartas.items():
            if minha_mao in v:
                x = k

    return (x)

estagio_jogo_antigo = 0
estagio_jogo = 1 #apenas para inicializar diferente de estagio_jogo
num_oponentes_antigos = 10
chance_de_ganhar = 0
chance_de_empatar = 0
reducao_chance_de_ganhar = 0.15
# f = open(filepath + "/log.txt", 'a')

wb = openpyxl.Workbook()
excel = wb.active
#cabeçalho do excel
excel.cell(row = 1, column = 1).value = 'id_contador' 
excel.cell(row = 1, column = 2).value = 'data_hora1'
excel.cell(row = 1, column = 3).value = 'qtdd_call'
excel.cell(row = 1, column = 4).value = 'qtdd_pot'
excel.cell(row = 1, column = 5).value = 'estagio_jogo'
excel.cell(row = 1, column = 6).value = 'carta1'
excel.cell(row = 1, column = 7).value = 'carta2'
excel.cell(row = 1, column = 8).value = 'carta_mesa1'
excel.cell(row = 1, column = 9).value = 'carta_mesa2'
excel.cell(row = 1, column = 10).value = 'carta_mesa3'
excel.cell(row = 1, column = 11).value = 'carta_mesa4'
excel.cell(row = 1, column = 12).value = 'carta_mesa5'
excel.cell(row = 1, column = 13).value = 'reducao_chance_de_ganhar' 
excel.cell(row = 1, column = 14).value = 'chance_de_ganhar' 
excel.cell(row = 1, column = 15).value = 'chance_de_empatar' 
excel.cell(row = 1, column = 16).value = 'chance_de_perder' 
excel.cell(row = 1, column = 17).value = 'ev' 
excel.cell(row = 1, column = 18).value = 'posicao_oponentes' 
excel.cell(row = 1, column = 19).value = 'num_oponentes' 
excel.cell(row = 1, column = 20).value = 'pos_dealer' 
excel.cell(row = 1, column = 21).value = 'meu_stack' 
excel.cell(row = 1, column = 22).value = 'stack_oponente_2' 
excel.cell(row = 1, column = 23).value = 'stack_oponente_3' 
excel.cell(row = 1, column = 24).value = 'stack_oponente_4' 
excel.cell(row = 1, column = 25).value = 'stack_oponente_5' 
excel.cell(row = 1, column = 26).value = 'stack_oponente_6' 
excel.cell(row = 1, column = 27).value = 'stack_oponente_7' 
excel.cell(row = 1, column = 28).value = 'stack_oponente_8' 
excel.cell(row = 1, column = 29).value = 'stack_oponente_9' 
excel.cell(row = 1, column = 30).value = 'numero_classificacao_mao' 
excel.cell(row = 1, column = 31).value = 'numero_acao' 
excel.cell(row = 1, column = 32).value = 'acao_tomada'
#tudo do bb agora
excel.cell(row = 1, column = 33).value = 'bb'
excel.cell(row = 1, column = 34).value = 'qtdd_call_bb'
excel.cell(row = 1, column = 35).value = 'qtdd_pot_bb'
excel.cell(row = 1, column = 36).value = 'ev_bb'
excel.cell(row = 1, column = 37).value = 'meu_stack_bb'
excel.cell(row = 1, column = 38).value = 'stack_oponente_2_bb'
excel.cell(row = 1, column = 39).value = 'stack_oponente_3_bb'
excel.cell(row = 1, column = 40).value = 'stack_oponente_4_bb'
excel.cell(row = 1, column = 41).value = 'stack_oponente_5_bb'
excel.cell(row = 1, column = 42).value = 'stack_oponente_6_bb'
excel.cell(row = 1, column = 43).value = 'stack_oponente_7_bb'
excel.cell(row = 1, column = 44).value = 'stack_oponente_8_bb'
excel.cell(row = 1, column = 45).value = 'stack_oponente_9_bb'
excel.cell(row = 1, column = 46).value = 'aposta_jog_2'
excel.cell(row = 1, column = 47).value = 'aposta_jog_3'
excel.cell(row = 1, column = 48).value = 'aposta_jog_4'
excel.cell(row = 1, column = 49).value = 'aposta_jog_5'
excel.cell(row = 1, column = 50).value = 'aposta_jog_6'
excel.cell(row = 1, column = 51).value = 'aposta_jog_2_bb'
excel.cell(row = 1, column = 52).value = 'aposta_jog_3_bb'
excel.cell(row = 1, column = 53).value = 'aposta_jog_4_bb'
excel.cell(row = 1, column = 54).value = 'aposta_jog_5_bb'
excel.cell(row = 1, column = 55).value = 'aposta_jog_6_bb'
excel.cell(row = 1, column = 56).value = 'lista_oponentes_com_aposta'
excel.cell(row = 1, column = 57).value = 'id_mao'
excel.cell(row = 1, column = 58).value = 'id_sequencia'
excel.cell(row = 1, column = 59).value = 'id_sequencia_etapa'



wb.save(filepath+"\log.xlsx")

linha = 2
id_contador = 1
id_mao = 0
id_sequencia = 1
id_sequencia_etapa = 1
carta1_antiga = 0
carta2_antiga = 0
estagio_jogo_antigo2 = 0

lista_duracao = [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,
                 1,1,1,1,1,1,1,1,1,1,1,1,
                 2,2,2,2,
                 3,3,3,3,
                 4,4,4,4,
                 5,5,5,5,
                 6,6,6,6,
                 7,7,7,7,
                 8,8,8,8,
                 9,9,9,9,
                 10,10,10,
                 11,11,11,
                 12,12,12,
                 13,13,13,
                 14,14,14,
                 15,15,15,
                 16,16,16,
                 17,17,17,
                 18,18,18,
                 19,19,19,
                 20,20,
                 21,21,
                 22,22,
                 23,23,
                 24,24,
                 25,25]
tamanho_lista_duracao = len(lista_duracao)

# --------------------------------------------------------------------------------------------------------------------------------------------------------------
# ------------------ PROGRAMA COMEÇA AQUI ----------------------------------------------------------------------------------------------------------------------
# --------------------------------------------------------------------------------------------------------------------------------------------------------------


print('digite o bb: \n')
bb = float(input())

while keyboard.is_pressed('q') == False:
    data_hora1 = strftime("%y_%m_%d %H_%M_%S")
    fazer_screenshot(data_hora1, filepath)
    image = Image.open(filepath + '/' +  data_hora1 + ".jpg")

    acao = fazer_acao(image)

    if acao == 0:
        pass
    else:
        numero_randomico = random.randint(0,tamanho_lista_duracao-1)
        sleep(lista_duracao[numero_randomico]/10)

        image_carta1 = image.crop((minha_carta_1[0], minha_carta_1[1], minha_carta_1[2], minha_carta_1[3]))
        image_carta2 = image.crop((minha_carta_2[0], minha_carta_2[1], minha_carta_2[2], minha_carta_2[3]))

        image_carta_mesa1 = image.crop((carta_mesa_1[0], carta_mesa_1[1], carta_mesa_1[2], carta_mesa_1[3]))
        image_carta_mesa2 = image.crop((carta_mesa_2[0], carta_mesa_2[1], carta_mesa_2[2], carta_mesa_2[3]))
        image_carta_mesa3 = image.crop((carta_mesa_3[0], carta_mesa_3[1], carta_mesa_3[2], carta_mesa_3[3]))
        image_carta_mesa4 = image.crop((carta_mesa_4[0], carta_mesa_4[1], carta_mesa_4[2], carta_mesa_4[3]))
        image_carta_mesa5 = image.crop((carta_mesa_5[0], carta_mesa_5[1], carta_mesa_5[2], carta_mesa_5[3]))
        
        image_carta1_naipe = image.crop((minha_carta_1_naipe[0], minha_carta_1_naipe[1], minha_carta_1_naipe[2], minha_carta_1_naipe[3]))
        image_carta2_naipe = image.crop((minha_carta_2_naipe[0], minha_carta_2_naipe[1], minha_carta_2_naipe[2], minha_carta_2_naipe[3]))

        image_carta_mesa1_naipe = image.crop((carta_mesa_1_naipe[0], carta_mesa_1_naipe[1], carta_mesa_1_naipe[2], carta_mesa_1_naipe[3]))
        image_carta_mesa2_naipe = image.crop((carta_mesa_2_naipe[0], carta_mesa_2_naipe[1], carta_mesa_2_naipe[2], carta_mesa_2_naipe[3]))
        image_carta_mesa3_naipe = image.crop((carta_mesa_3_naipe[0], carta_mesa_3_naipe[1], carta_mesa_3_naipe[2], carta_mesa_3_naipe[3]))
        image_carta_mesa4_naipe = image.crop((carta_mesa_4_naipe[0], carta_mesa_4_naipe[1], carta_mesa_4_naipe[2], carta_mesa_4_naipe[3]))
        image_carta_mesa5_naipe = image.crop((carta_mesa_5_naipe[0], carta_mesa_5_naipe[1], carta_mesa_5_naipe[2], carta_mesa_5_naipe[3]))

        try:
            carta1 = (reconhecer_carta(image_carta1) + reconhecer_carta_naipe(image_carta1_naipe))
        except:
            carta1 = 0
        try:
            carta2 = (reconhecer_carta(image_carta2) + reconhecer_carta_naipe(image_carta2_naipe))
        except:
            carta2 = 0

        try:
            carta_mesa1 = (reconhecer_carta(image_carta_mesa1) + reconhecer_carta_naipe(image_carta_mesa1_naipe))
        except:
            carta_mesa1 = 0
        try:
            carta_mesa2 = (reconhecer_carta(image_carta_mesa2) + reconhecer_carta_naipe(image_carta_mesa2_naipe))
        except:
            carta_mesa2 = 0
        try:
            carta_mesa3 = (reconhecer_carta(image_carta_mesa3) + reconhecer_carta_naipe(image_carta_mesa3_naipe))
        except:
            carta_mesa3 = 0
        try:
            carta_mesa4 = (reconhecer_carta(image_carta_mesa4) + reconhecer_carta_naipe(image_carta_mesa4_naipe))
        except:
            carta_mesa4 = 0
        try:
            carta_mesa5 = (reconhecer_carta(image_carta_mesa5) + reconhecer_carta_naipe(image_carta_mesa5_naipe))
        except:
            carta_mesa5 = 0

        estagio_jogo = reconhecer_estagio_jogo(carta1, carta2, carta_mesa1, carta_mesa2, carta_mesa3, carta_mesa4, carta_mesa5)

        posicao_oponentes , num_oponentes = reconhecer_numero_oponentes(image)

    
        if image.getpixel((1305, 992)) == (191, 35, 36):
            qtdd_call = reconhecer_numero(image, (1290, 985, 1582, 1063))
        else:
            qtdd_call = 0
    
        qtdd_pot = reconhecer_numero(image, (890 , 326, 1032, 365))



        #acho que posso tirar esses ifs, já que coloquei dentro do acao <> 0
        if estagio_jogo == 'sem cartas':
            pass    
        elif (estagio_jogo_antigo == estagio_jogo) and (num_oponentes_antigos == num_oponentes): # mesmo estagio de jogo e mesmo numero oponentes
            pass
        elif carta1 == 0 or carta2 == 0:
            pass
        else:
            chance_de_ganhar, chance_de_empatar = calcular_chance_de_ganhar(carta1, carta2, carta_mesa1, carta_mesa2, carta_mesa3, carta_mesa4, carta_mesa5, acao, num_oponentes)
            estagio_jogo_antigo = estagio_jogo
            num_oponentes_antigos = num_oponentes

        if chance_de_ganhar <= reducao_chance_de_ganhar:
            chance_de_ganhar = 0
        else:
            chance_de_ganhar = chance_de_ganhar - reducao_chance_de_ganhar

        pos_dealer = reconhecer_dealer(image)

        meu_stack = reconhecer_numero(image, (952, 795, 1041, 840))

        stack_oponente_2 = reconhecer_numero(image, (300, 600, 390, 640))
        stack_oponente_3 = reconhecer_numero(image, (351, 270, 437, 315))
        stack_oponente_4 = reconhecer_numero(image, (875, 173, 979, 212))
        stack_oponente_5 = reconhecer_numero(image, (1485, 270, 1573, 318))
        stack_oponente_6 = reconhecer_numero(image, (1535, 605, 1625, 637))
        # stack_oponente_7 = reconhecer_numero(image, (1462, 274, 1601, 313))
        # stack_oponente_8 = reconhecer_numero(image, (1523, 484, 1667, 519))
        # stack_oponente_9 = reconhecer_numero(image, (1370, 701, 1517, 735))
        stack_oponente_7 = 0
        stack_oponente_8 = 0
        stack_oponente_9 = 0
    
        aposta_jog_2 = reconhecer_numero(image, (585, 561, 657, 595))
        aposta_jog_3 = reconhecer_numero(image, (599, 330, 682, 373))
        aposta_jog_4 = reconhecer_numero(image, (1030, 233, 1116, 270))
        aposta_jog_5 = reconhecer_numero(image, (1214, 289, 1297, 335))
        aposta_jog_6 = reconhecer_numero(image, (1257, 561, 1333, 595))

        minha_mao = [carta1, carta2]
        numero_classificacao_mao = classificar_mao(minha_mao)

        lista_oponentes_com_aposta = lista_apostas_oponentes(image)


    if estagio_jogo == 'sem cartas':
            # print('Hora: ' + filepath + '/' +  data_hora1 + ".jpg")
            # print('estagio do jogo eh: ', estagio_jogo)
            # f.write('\nHora: ' + filepath + '/' + data_hora1 + ".jpg")
            # f.write('\nestagio do jogo eh: '+ estagio_jogo)
            os.remove(filepath + '/' +  data_hora1 + ".jpg")
    elif acao == 0:
            # print('Hora: ' + filepath + '/' +  data_hora1 + ".jpg")
            # print('numero acao eh: ', acao)
            # f.write('\nHora: ' + filepath + '/' + data_hora1 + ".jpg")
            # f.write('\no numero acao eh: '+ str(acao))
            os.remove(filepath + '/' +  data_hora1 + ".jpg")
    else:
        
        if carta1 == carta1_antiga and carta2 == carta2_antiga:
            if estagio_jogo_antigo2 == estagio_jogo:
                id_sequencia = id_sequencia + 1
                id_sequencia_etapa = id_sequencia_etapa + 1
            else:
                estagio_jogo_antigo2 = estagio_jogo
                id_sequencia = id_sequencia + 1
                id_sequencia_etapa = 1
        else:
            id_sequencia = 1
            id_mao = id_mao + 1
            carta1_antiga = carta1
            carta2_antiga = carta2
            id_sequencia_etapa = 1
            estagio_jogo_antigo2 = 0
        
        # print('Hora: ' + filepath + '/' +  data_hora1 + ".jpg")
        # print('qtdd para call eh: ', qtdd_call)
        # print('qtdd do pot eh: ', qtdd_pot)
        # print('estagio do jogo eh: ', estagio_jogo)
        # print('cartas da mesa sao: ', carta1, carta2, carta_mesa1, carta_mesa2, carta_mesa3, carta_mesa4, carta_mesa5)
        # print('o numero acao eh: ', acao)
        print('reducao chance de ganhar eh: ', reducao_chance_de_ganhar)
        print('chance de ganhar: ', chance_de_ganhar)
        print('chance de empatar eh: ', chance_de_empatar)
        chance_de_perder = 1-chance_de_ganhar-chance_de_empatar
        print('chance de perder eh: ', chance_de_perder)
        ev = chance_de_ganhar * (qtdd_pot + qtdd_call) + chance_de_empatar * (qtdd_pot + qtdd_call)/2 - (1-chance_de_ganhar-chance_de_empatar) * qtdd_call
        print('ev eh: ', ev)
        # print('posicao oponentes: ', posicao_oponentes)
        # print('numero de oponente: ', num_oponentes)
        # print('posicao dealer eh: ', pos_dealer)
        # print('meu stack eh: ', meu_stack)
        # print('stacks: op.2 = ', stack_oponente_2, ' op.3 = ', stack_oponente_3, ' op.4 = ', stack_oponente_4, ' op.5 = ', stack_oponente_5, ' op.6 = ', stack_oponente_6, ' op.7 = ', stack_oponente_7, ' op.8 = ', stack_oponente_8, ' op.9 = ', stack_oponente_9)
        print('forca da mao: ' + str(numero_classificacao_mao))
        print('apostas: op.2= ', aposta_jog_2, ' op.3= ', aposta_jog_3, ' op.4= ', aposta_jog_4, ' op.5= ', aposta_jog_5, ' op.6= ', aposta_jog_6)
        print('lista_aposta_oponentes= ', lista_oponentes_com_aposta)
        print('id mao: ', id_mao)
        print('id sequencia: ',id_sequencia)
        print('id sequencia mesa: ', id_sequencia_etapa)
        # print('carta antiga 1 e 2: ', carta1_antiga, ' ', carta2_antiga)
        print('/----------------------------------------------------------------------------------------------------/')



        #mudança das variaveis para bb
        qtdd_call_bb = qtdd_call / bb
        qtdd_pot_bb = qtdd_pot / bb
        ev_bb = ev / bb
        meu_stack_bb = meu_stack / bb
        stack_oponente_2_bb = stack_oponente_2 / bb
        stack_oponente_3_bb = stack_oponente_3 / bb
        stack_oponente_4_bb = stack_oponente_4 / bb
        stack_oponente_5_bb = stack_oponente_5 / bb
        stack_oponente_6_bb = stack_oponente_6 / bb
        stack_oponente_7_bb = stack_oponente_7 / bb
        stack_oponente_8_bb = stack_oponente_8 / bb
        stack_oponente_9_bb = stack_oponente_9 / bb
        aposta_jog_2_bb = aposta_jog_2 / bb
        aposta_jog_3_bb = aposta_jog_3 / bb
        aposta_jog_4_bb = aposta_jog_4 / bb
        aposta_jog_5_bb = aposta_jog_5 / bb
        aposta_jog_6_bb = aposta_jog_6 / bb

        #retirado pois uso somente o excel
        # f.write('\nHora: ' + str(filepath) + '/' +  str(data_hora1) + ".jpg")
        # f.write('\nqtdd para call eh: '+ str(qtdd_call))
        # f.write('\nqtdd do pot eh: '+ str(qtdd_pot))
        # f.write('\nestagio do jogo eh: '+ estagio_jogo)
        # f.write('\ncartas da mesa sao: '+ str(carta1)+ ' ' +str(carta2)+ ' ' + str(carta_mesa1)+ ' ' + str(carta_mesa2)+ ' ' + str(carta_mesa3)+ ' ' + str(carta_mesa4)+ ' ' + str(carta_mesa5))
        # f.write('\no numero acao eh: '+ str(acao))
        # f.write('\nreducao chance de ganhar eh: '+ str(round(reducao_chance_de_ganhar*100,1)))
        # f.write('\nchance de ganhar: '+ str(round(chance_de_ganhar*100,1)))
        # f.write('\nchance de empatar eh: '+ str(round(chance_de_empatar*100,1)))
        # f.write('\nchance de perder eh: '+ str(round(chance_de_perder*100,1)))
        # f.write('\nev eh: '+ str(ev))
        # f.write('\nposicao oponentes: '+ str(posicao_oponentes))
        # f.write('\nnumero de oponente: '+ str(num_oponentes))
        # f.write('\nposicao dealer eh: '+ str(pos_dealer))
        # f.write('\nmeu stack eh: '+ str(meu_stack))
        # f.write('\nstacks: op.2 = '+ str(stack_oponente_2)+ ' op.3 = '+ str(stack_oponente_3)+ ' op.4 = '+ str(stack_oponente_4)+ ' op.5 = '+ str(stack_oponente_5)+ ' op.6 = '+ str(stack_oponente_6)+ ' op.7 = '+ str(stack_oponente_7)+ ' op.8 = '+ str(stack_oponente_8)+ ' op.9 = '+ str(stack_oponente_9))
        # f.write('\nforca da mao: ' + str(numero_classificacao_mao))
        # f.write('\nbb eh: '+str(bb))
        # f.write('\nqtdd_call_bb eh: '+str(qtdd_call_bb))
        # f.write('\nqtdd_pot_bb eh: '+str(qtdd_pot_bb))
        # f.write('\nev_bb eh: '+str(ev_bb))
        # f.write('\nmeu_stack_bb eh: '+str(meu_stack_bb))
        # f.write('\nstack_oponente_2_bb eh: '+str(stack_oponente_2_bb))
        # f.write('\nstack_oponente_3_bb eh: '+str(stack_oponente_3_bb))
        # f.write('\nstack_oponente_4_bb eh: '+str(stack_oponente_4_bb))
        # f.write('\nstack_oponente_5_bb eh: '+str(stack_oponente_5_bb))
        # f.write('\nstack_oponente_6_bb eh: '+str(stack_oponente_6_bb))
        # f.write('\nstack_oponente_7_bb eh: '+str(stack_oponente_7_bb))
        # f.write('\nstack_oponente_8_bb eh: '+str(stack_oponente_8_bb))
        # f.write('\nstack_oponente_9_bb eh: '+str(stack_oponente_9_bb))
        
        # f.write('\naposta_jog_2_bb eh: '+str(aposta_jog_2_bb))
        # f.write('\naposta_jog_3_bb eh: '+str(aposta_jog_3_bb))
        # f.write('\naposta_jog_4_bb eh: '+str(aposta_jog_4_bb))
        # f.write('\naposta_jog_5_bb eh: '+str(aposta_jog_5_bb))
        # f.write('\naposta_jog_6_bb eh: '+str(aposta_jog_6_bb))


        excel.cell(row = linha, column = 1).value = id_contador 
        excel.cell(row = linha, column = 2).value = data_hora1
        excel.cell(row = linha, column = 3).value = qtdd_call
        excel.cell(row = linha, column = 4).value = qtdd_pot
        excel.cell(row = linha, column = 5).value = estagio_jogo
        excel.cell(row = linha, column = 6).value = carta1
        excel.cell(row = linha, column = 7).value = carta2
        excel.cell(row = linha, column = 8).value = carta_mesa1
        excel.cell(row = linha, column = 9).value = carta_mesa2
        excel.cell(row = linha, column = 10).value = carta_mesa3
        excel.cell(row = linha, column = 11).value = carta_mesa4
        excel.cell(row = linha, column = 12).value = carta_mesa5
        excel.cell(row = linha, column = 13).value = reducao_chance_de_ganhar 
        excel.cell(row = linha, column = 14).value = chance_de_ganhar 
        excel.cell(row = linha, column = 15).value = chance_de_empatar 
        excel.cell(row = linha, column = 16).value = chance_de_perder 
        excel.cell(row = linha, column = 17).value = ev 
        excel.cell(row = linha, column = 18).value = str(posicao_oponentes) 
        excel.cell(row = linha, column = 19).value = num_oponentes 
        excel.cell(row = linha, column = 20).value = pos_dealer 
        excel.cell(row = linha, column = 21).value = meu_stack 
        excel.cell(row = linha, column = 22).value = stack_oponente_2 
        excel.cell(row = linha, column = 23).value = stack_oponente_3 
        excel.cell(row = linha, column = 24).value = stack_oponente_4 
        excel.cell(row = linha, column = 25).value = stack_oponente_5 
        excel.cell(row = linha, column = 26).value = stack_oponente_6 
        excel.cell(row = linha, column = 27).value = stack_oponente_7 
        excel.cell(row = linha, column = 28).value = stack_oponente_8 
        excel.cell(row = linha, column = 29).value = stack_oponente_9 
        excel.cell(row = linha, column = 30).value = numero_classificacao_mao 
        excel.cell(row = linha, column = 31).value = acao 
        #excel.cell(row = linha, column = 32).value = 'acao_tomada' (essa é feita diretamente na função)
        excel.cell(row = linha, column = 33).value = bb 
        excel.cell(row = linha, column = 34).value = qtdd_call_bb 
        excel.cell(row = linha, column = 35).value = qtdd_pot_bb 
        excel.cell(row = linha, column = 36).value = ev_bb 
        excel.cell(row = linha, column = 37).value = meu_stack_bb 
        excel.cell(row = linha, column = 38).value = stack_oponente_2_bb 
        excel.cell(row = linha, column = 39).value = stack_oponente_3_bb
        excel.cell(row = linha, column = 40).value = stack_oponente_4_bb
        excel.cell(row = linha, column = 41).value = stack_oponente_5_bb
        excel.cell(row = linha, column = 42).value = stack_oponente_6_bb
        excel.cell(row = linha, column = 43).value = stack_oponente_7_bb
        excel.cell(row = linha, column = 44).value = stack_oponente_8_bb
        excel.cell(row = linha, column = 45).value = stack_oponente_9_bb
        excel.cell(row = linha, column = 46).value = aposta_jog_2
        excel.cell(row = linha, column = 47).value = aposta_jog_3
        excel.cell(row = linha, column = 48).value = aposta_jog_4
        excel.cell(row = linha, column = 49).value = aposta_jog_5
        excel.cell(row = linha, column = 50).value = aposta_jog_6
        excel.cell(row = linha, column = 51).value = aposta_jog_2_bb
        excel.cell(row = linha, column = 52).value = aposta_jog_3_bb
        excel.cell(row = linha, column = 53).value = aposta_jog_4_bb
        excel.cell(row = linha, column = 54).value = aposta_jog_5_bb
        excel.cell(row = linha, column = 55).value = aposta_jog_6_bb
        excel.cell(row = linha, column = 56).value = str(lista_oponentes_com_aposta)
        excel.cell(row = linha, column = 57).value = id_mao
        excel.cell(row = linha, column = 58).value = id_sequencia
        excel.cell(row = linha, column = 59).value = id_sequencia_etapa


        wb.save(filepath+"\log.xlsx")

        linha = linha + 1
        id_contador = id_contador + 1


    if acao == 0:
        pass
    elif estagio_jogo == 'pre-flop':
        if acao == 3: #posso fazer mesa ou desistir
            if numero_classificacao_mao < 20:
                mesa()
                # centralizar_mouse()
            else:
                desistir()
                # centralizar_mouse()
        elif acao == 5: #posso apostar ou desistir
            if numero_classificacao_mao < 20:
                apostar()
                # centralizar_mouse()
            else:
                desistir()
                # centralizar_mouse()
        elif acao == 6: #posso apostar ou mesa
            if numero_classificacao_mao < 10:
                apostar()
                # centralizar_mouse()
            else:
                mesa()
                # centralizar_mouse()
        else: #posso apostar, mesa e desistir
            if qtdd_call > 0:
                if numero_classificacao_mao < 10:
                    apostar()
                    # centralizar_mouse()
                elif numero_classificacao_mao < 25:
                    mesa()
                    # centralizar_mouse()
                else:
                    desistir()
                    # centralizar_mouse()
            else: #ninguem fez aposta ainda
                if numero_classificacao_mao < 15:
                    apostar()
                    # centralizar_mouse()
                else:
                    mesa()

    #aqui estamos pelo menos no flop. retirar call e aposta quando chance de ganhar = 0
    elif acao == 3: #posso fazer mesa ou desistir
        if ev > 0:
            if chance_de_ganhar > 0:         
                mesa()
                # centralizar_mouse()
        else:
            desistir()
            # centralizar_mouse()
    elif acao == 5: #posso apostar ou desistir
        if ev > 0:
            apostar()
            # centralizar_mouse()
        else:
            desistir()
            # centralizar_mouse()
    elif acao == 6: #posso apostar ou mesa
        if ev > 0:
            if chance_de_ganhar > 0.5:
                apostar()
                # centralizar_mouse()
            else:
                mesa()
                # centralizar_mouse()
        else:
            mesa()
            # centralizar_mouse()
    else: #posso apostar, mesa e desistir
        if ev > 0:
            if chance_de_ganhar > 0.5:
                    apostar()
                    # centralizar_mouse()
            else:
                mesa()
                # centralizar_mouse()
        else:
            if qtdd_call == 0:
                mesa()
                # centralizar_mouse()
            else:
                desistir()
                # centralizar_mouse()


    # print('/----------------------------------------------------------------------------------------------------/')
    # f.write('\n/----------------------------------------------------------------------------------------------------/'):

    sleep(1)
